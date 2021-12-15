# This is a script to generate grading rubrics to help you determine your final grade in
# any of your courses that do not update grades online regularly. Requires the openpyxl
# library to be installed in your current python environment, and also requires
# Python 3.6+ because of it's use of f-strings.
# AUTHOR: Robert Drescher

import openpyxl


def main():
    print("This script will run you through the process of making a grading rubric for any of your classes that you can open in Excel/Numbers/etc.\n\n")

    ###########################################################################################
    ################################### OUTPUT OPTIONS ########################################
    ###########################################################################################
    POINTS = False
    TOTAL_POINTS = 0
    HEADERS = True
    STARTING_ROW = 2
    HEADER_ROW = 1

    FINAL_PERCENTAGE_ROW = 5
    FINAL_PERCENTAGE_COLUMN = 7

    FINAL_LETTER_ROW = 10
    FINAL_LETTER_COLUMN = 7

    ASSIGNMENT_NAME_COLUMN = 2
    ASSIGNMENT_WEIGHT_COLUMN = 3
    ASSIGNMENT_GRADE_COLUMN = 4

    LETTER_GRADE_LETTER_COLUMN = 9
    LETTER_GRADE_MINIMUM_COLUMN = 10


    ASSIGNMENT_WEIGHT_COL_LETTER = chr(ASSIGNMENT_WEIGHT_COLUMN + 64)
    ASSIGNMENT_GRADE_COL_LETTER = chr(ASSIGNMENT_GRADE_COLUMN + 64)


    ###########################################################################################
    ############################ Determine Points or Percentages ##############################
    ###########################################################################################
    while True:
        choice = input("Do you want to use points or percentages for grade calcualtion?\n1 for Points\n2 for Percentages: ")
        if choice.strip() == "1":
            POINTS = True
            break
        elif choice.strip() == "2":
            POINTS = False
            break
        else:
            print("Invalid input. Please try again.")



    ###########################################################################################
    ########################### Getting Letters for Grades ####################################
    ###########################################################################################
    # LETTER_GRADES = ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D'] #This line works for my school but the below code will work for all schools...
    LETTER_GRADES = []

    while True:
        letter_grade = input("Input a grade value or press enter to stop: ")

        if letter_grade:
            LETTER_GRADES.append(letter_grade)
        else: break


    LETTER_GRADES_DICT = {} #ORDER NOT GUARANTEED, USE A SORTED LIST OF KEYS
    ASSIGNMENTS_DICT = {} #ORDER NOT GUARANTEED, USE A SORTED LIST OF KEYS



    ###########################################################################################
    #################### Getting minimum values needed for grades #############################
    ###########################################################################################
    for letter in LETTER_GRADES:
        def askMinValue(letter):
            minimum = input(f"What is the minimum grade you need to get a(n) {letter}? ")
            try:
                minimum = int(minimum)
            except ValueError:
                print(f"INVALID VALUE GIVEN, PLEASE TRY AGAIN")
                minimum = askMinValue(letter)
            return minimum
        minimum = askMinValue(letter)
        LETTER_GRADES_DICT[letter] = minimum

    print("\n\n\n\n")



    ###########################################################################################
    ################## Getting Assignments and their weights ##################################
    ###########################################################################################
    print("Enter assignment names and weights, leave name blank to end.")
    if POINTS:
        print("Weights should be the number of poitns the asignment is worth.")
    else:
        print("Weights should be a decimal, (e.g. a paper worth 10% of your grade would have a decimal weight of 0.10)")
    while True:
        assignmentName = input("What is the name of the assignment? ")

        if assignmentName:
            if POINTS:
                points = input("How many points is the assignment worth? ")
                try:
                    points = int(points)
                    if points:
                        ASSIGNMENTS_DICT[assignmentName] = points
                except ValueError:
                    print("Invalid value given, please try again.")
                    continue
                ASSIGNMENTS_DICT[assignmentName] = points
                TOTAL_POINTS += points
            else:
                weight = input('What is the weight of this assignment? ')
                try:
                    weight = float(weight)
                    if weight > 1.0 or weight < 0.0:
                        raise ValueError
                except ValueError:
                    print("Invalid input given for weight, please try again.")
                    continue
                ASSIGNMENTS_DICT[assignmentName] = weight
        else: break

    print('\n\n\n\n')



    ###########################################################################################
    ################################ Create excel sheet #######################################
    ###########################################################################################
    wbook = openpyxl.Workbook()
    sheet = wbook[wbook.sheetnames[0]]



    ###########################################################################################
    ################################# Letters and mins ########################################
    ###########################################################################################
    row = STARTING_ROW
    for letterGrade in sorted(LETTER_GRADES_DICT.keys()):
        letter = letterGrade
        minimum = LETTER_GRADES_DICT[letterGrade]

        letter_cell = sheet.cell(row, LETTER_GRADE_LETTER_COLUMN)
        minimum_cell = sheet.cell(row, LETTER_GRADE_MINIMUM_COLUMN)

        letter_cell.value = letter
        minimum_cell.value = minimum
        LETTER_GRADES_DICT[letterGrade] = minimum_cell.coordinate
        
        row += 1



    ###########################################################################################
    ############################# Assignments and weights #####################################
    ###########################################################################################
    row = STARTING_ROW
    for assignmentName in sorted(ASSIGNMENTS_DICT.keys(), key=lambda key: ASSIGNMENTS_DICT[key]):
        weight = ASSIGNMENTS_DICT[assignmentName]

        name_cell = sheet.cell(row, ASSIGNMENT_NAME_COLUMN)
        weight_cell = sheet.cell(row, ASSIGNMENT_WEIGHT_COLUMN)
        grade_cell = sheet.cell(row, ASSIGNMENT_GRADE_COLUMN)

        name_cell.value = assignmentName
        weight_cell.value = weight
        grade_cell.value = 100 if not POINTS else weight

        row += 1



    ###########################################################################################
    ################################## Final Percentage #######################################
    ###########################################################################################
    MAX_ASSIGNMENT_ROW = row

    final_percentage_formula = ""

    if POINTS:
        # Can just be SUM(D:D) / SUM(C:C)
        final_percentage_formula = f"=SUM({ASSIGNMENT_GRADE_COL_LETTER}:{ASSIGNMENT_GRADE_COL_LETTER}) / SUM({ASSIGNMENT_WEIGHT_COL_LETTER}:{ASSIGNMENT_WEIGHT_COL_LETTER})*100"
    else:
        #Final formula needs to look similar to "=SUM(A1*B2, A2*B2, ...)"
        final_percentage_formula = "=SUM(" 
        first = True
        for row in range(STARTING_ROW, MAX_ASSIGNMENT_ROW):
            weight_cell = sheet.cell(row, ASSIGNMENT_WEIGHT_COLUMN)
            grade_cell = sheet.cell(row, ASSIGNMENT_GRADE_COLUMN)

            weight_cell_coord = weight_cell.coordinate
            grade_cell_coord = grade_cell.coordinate
            if first:
                final_percentage_formula += f'{weight_cell_coord}*{grade_cell_coord}'
                first = not first
            else:
                final_percentage_formula += f',{weight_cell_coord}*{grade_cell_coord}'

        final_percentage_formula += ")"

    FINAL_PERCENTAGE_CELL = sheet.cell(FINAL_PERCENTAGE_ROW, FINAL_PERCENTAGE_COLUMN)
    FINAL_PERCENTAGE_CELL.value = final_percentage_formula

    if HEADERS:
        cell = sheet.cell(FINAL_PERCENTAGE_ROW, FINAL_PERCENTAGE_COLUMN - 1)
        cell.value = "Final Percentage"



    ###########################################################################################
    #################################### Final Letter #########################################
    ###########################################################################################
    FINAL_LETTER_CELL = sheet.cell(FINAL_LETTER_ROW, FINAL_LETTER_COLUMN)

    #Final formula needs to look similar to "=IF(A1>F1, A, IF(A1>F2, B, IF(A1>F3, C, ....))"
    final_letter_formula = f"=IF("
    paren_counter = 0
    for letter in sorted(LETTER_GRADES_DICT):
        minval = LETTER_GRADES_DICT[letter]
        final_letter_formula += f'{FINAL_PERCENTAGE_CELL.coordinate} >= {minval}, "{letter}", IF('
        paren_counter += 1

    final_letter_formula = final_letter_formula[0:len(final_letter_formula)-5]
    final_letter_formula += (")"*paren_counter)
    FINAL_LETTER_CELL.value = final_letter_formula



    ###########################################################################################
    ######################################## Headers ##########################################
    ###########################################################################################
    if HEADERS:
        cell = sheet.cell(FINAL_PERCENTAGE_ROW, FINAL_PERCENTAGE_COLUMN - 1)
        cell.value = "Final Percentage"

        cell = sheet.cell(FINAL_LETTER_ROW, FINAL_LETTER_COLUMN - 1)
        cell.value = "Final Letter Grade"

        cell = sheet.cell(HEADER_ROW, ASSIGNMENT_NAME_COLUMN)
        cell.value = "Assignment Name"

        cell = sheet.cell(HEADER_ROW, ASSIGNMENT_WEIGHT_COLUMN)
        cell.value = "Weight" if not POINTS else "Point Value"
        
        cell = sheet.cell(HEADER_ROW, ASSIGNMENT_GRADE_COLUMN)
        cell.value = "Grade" if not POINTS else "Points Earned"

        cell = sheet.cell(HEADER_ROW, LETTER_GRADE_LETTER_COLUMN)
        cell.value = "Letter Grade"

        cell = sheet.cell(HEADER_ROW, LETTER_GRADE_MINIMUM_COLUMN)
        cell.value = "Minimum Percentage"



    ###########################################################################################
    #################################### Output Excel File ####################################
    ###########################################################################################
    outputName = input("Enter a name for your rubric file: ")
    if not outputName.endswith(".xlsx"): outputName += ".xlsx"
    wbook.save(outputName)

if __name__ == "__main__":
    main()