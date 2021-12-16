from tkinter import Tk, Button, Entry, Text, Frame, END
from tkinter.messagebox import showerror, showinfo, askyesno
from enum import Enum
import openpyxl

HEADERS = True
STARTING_ROW = 2
HEADER_ROW = 1

FINAL_PERCENTAGE_ROW = 5
FINAL_PERCENTAGE_COLUMN = 7

FINAL_LETTER_ROW = 10
FINAL_LETTER_COLUMN = 7

ASSIGNMENT_NAME_COLUMN = 2
ASSIGNMENT_WEIGHT_COLUMN = 3
ASSIGNMENT_GRADE_COLUMN = 4

LETTER_GRADE_LETTER_COLUMN = 9
LETTER_GRADE_MINIMUM_COLUMN = 10


ASSIGNMENT_WEIGHT_COL_LETTER = chr(ASSIGNMENT_WEIGHT_COLUMN + 64)
ASSIGNMENT_GRADE_COL_LETTER = chr(ASSIGNMENT_GRADE_COLUMN + 64)

class Step(Enum):
    START = 1
    GRADES = 2
    GRADE_MINIMUMS = 3
    ASSIGNMENTS = 4
    SUBMIT = 5

class RubricCreator(Frame):
    def __init__(self, parent, *args, **kwargs):
        Frame.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        self.bind("<Return>", lambda: self.next())

        self.POINTS = False
        
        self.LETTER_GRADES = []
        self.LAST_LETTER_GRADE = None
        self.LETTER_GRADES_DICT = {}
        self.ASSIGNMENTS_DICT = {}
        self.CURRENT_ASSIGNMENT_KEY = None
        self.CURRENT_ASSIGNMENT_WEIGHT = None

        self.step = Step.START
        self.step_flag = False #Determines whether this is the first time on this step.
        
        self.is_assignment_name = True

        self.text_area = Text(self)
        self.entry_area = Entry(self)
        self.next_button = Button(self, text="Next", command=lambda: self.next())

        self.text_area.pack()
        self.entry_area.pack(fill='both')
        self.next_button.pack()

    def progress_step(self):
        match self.step:
            case Step.START:
                self.step = Step.GRADES
                self.step_flag = False
            case Step.GRADES:
                self.step = Step.GRADE_MINIMUMS
                self.step_flag = False
            case Step.GRADE_MINIMUMS:
                self.step = Step.ASSIGNMENTS
                self.step_flag = False
            case Step.ASSIGNMENTS:
                self.step = Step.SUBMIT
                self.step_flag = False
            case _:
                showerror("Something Went Wrong", "Something went wrong. Please report this error on GitHub. ERROR CODE 12")

    def start(self):
        showinfo("Welcome", "This program will take you though creating a grading rubric. Enter your responses to the questions asked in the text box and hit Next after each one. Press OK to get started.")
        result = askyesno("Grading Method", "Do you want to use points for grading? Select 'No' to use percentage based grading and weights.")
        if result:
            self.POINTS = True
        self.step = Step.GRADES
        self.next()
    
    def get_entry_value(self) -> str:
        text = self.entry_area.get()
        return text
    
    def clear_entry_value(self):
        self.entry_area.delete(0, END)
    
    def write_to_text_area(self, string: str):
        self.text_area.delete(1.0, END)
        self.text_area.insert(1.0, string)

    def handle_grades(self):
        if self.step_flag: #Not the first run
            if self.get_entry_value():
                letter_grade = self.get_entry_value()
                self.clear_entry_value()
                self.LETTER_GRADES.append(letter_grade)
                self.write_to_text_area(f"Got {letter_grade}! Input the next grade you can get, leave blank to stop.")
            else: #Progress step if they left it blank
                self.LETTER_GRADES = self.LETTER_GRADES[::-1] #Reverse the list so it starts asking the next step with the first value input.
                self.progress_step()
        else: # is the first run
            self.write_to_text_area("Input a grade you can get, leave the entry blank to stop. (e.x. A)")
            self.step_flag = not self.step_flag

    def handle_grade_minimums(self):
        if self.step_flag:
            minimum = self.get_entry_value()
            try:
                minimum = int(minimum)
            except ValueError:
                showerror("You have to provide an integer.")
                return
            self.LETTER_GRADES_DICT[self.LAST_LETTER_GRADE] = minimum
            self.clear_entry_value()
            # Move to the next one.
            try:
                grade_val = self.LETTER_GRADES.pop()
                self.LAST_LETTER_GRADE = grade_val
                self.write_to_text_area(f"Input the minimum percentage you need to obtain a(n) {grade_val}")
            except IndexError:
                self.clear_entry_value()
                self.progress_step()
        else:
            grade_val = self.LETTER_GRADES.pop()
            self.LAST_LETTER_GRADE = grade_val
            self.write_to_text_area(f"Input the minimum percentage you need to obtain a(n) {grade_val}")
            self.step_flag = not self.step_flag
            
    def handle_assignments(self):
        if self.step_flag:
            if self.is_assignment_name:
                if self.get_entry_value():
                    self.CURRENT_ASSIGNMENT_KEY = self.get_entry_value()
                    self.write_to_text_area(f"How many points is {self.CURRENT_ASSIGNMENT_KEY} worth?")
                    self.is_assignment_name = not self.is_assignment_name
                    self.clear_entry_value()
                else: # End input
                    self.progress_step()
            else:
                if self.get_entry_value():
                    if self.POINTS:
                        try:
                            points = int(self.get_entry_value())
                            if points:
                                self.ASSIGNMENTS_DICT[self.CURRENT_ASSIGNMENT_KEY] = points
                                self.is_assignment_name = True
                                self.clear_entry_value()
                                self.write_to_text_area("Enter assignment names and weights. Leave the line blank twice to end. Weights should be the number of points the assignment is worth.\n\nWhat is the name of the assignment?")       
                        except ValueError:
                            showerror("Error", "Invalid points given, please try again.")
                            self.clear_entry_value()
                        
                    else:
                        try:
                            weight = float(self.get_entry_value())
                            if weight > 1.0 or weight < 0.0:
                                raise ValueError
                            self.ASSIGNMENTS_DICT[self.CURRENT_ASSIGNMENT_KEY] = weight
                            self.is_assignment_name = True
                            self.clear_entry_value()
                            self.write_to_text_area("Enter assignment names and weights. Weights should be a decimal (e.g. a paper worth 10% of your grade would have a decimal weight of 0.10) What is the name of the assignment?")
                        except ValueError:
                            showerror("Error", "Invalid weight given, please try again")
                            self.clear_entry_value()
                else:
                    showerror("Error", "You must input a weight for this assignment")
        else:
            if self.POINTS:
                self.write_to_text_area("Enter assignment names and weights. Leave the line blank twice to end. Weights should be the number of points the assignment is worth.\n\nWhat is the name of the assignment?")
            else:
                self.write_to_text_area("Enter assignment names and weights. Weights should be a decimal (e.g. a paper worth 10% of your grade would have a decimal weight of 0.10) What is the name of the assignment?")
            self.step_flag = not self.step_flag


    def handle_submit(self):
        if self.step_flag:
            output_name = self.get_entry_value()
            wbook = openpyxl.Workbook()
            sheet = wbook[wbook.sheetnames[0]]

            row = STARTING_ROW

            # Letters and Minimums
            for letter_grade in sorted(self.LETTER_GRADES_DICT.keys()):
                letter = letter_grade
                minimum = self.LETTER_GRADES_DICT[letter_grade]

                letter_cell = sheet.cell(row, LETTER_GRADE_LETTER_COLUMN)
                minimum_cell = sheet.cell(row, LETTER_GRADE_MINIMUM_COLUMN)

                letter_cell.value = letter
                minimum_cell.value = minimum
                self.LETTER_GRADES_DICT[letter_grade] = minimum_cell.coordinate

                row += 1
            
            # Assignments and Weights
            row = STARTING_ROW
            for assignment_name in sorted(self.ASSIGNMENTS_DICT.keys(), key=lambda key: self.ASSIGNMENTS_DICT[key]):
                weight = self.ASSIGNMENTS_DICT[assignment_name]

                name_cell = sheet.cell(row, ASSIGNMENT_NAME_COLUMN)
                weight_cell = sheet.cell(row, ASSIGNMENT_WEIGHT_COLUMN)
                grade_cell = sheet.cell(row, ASSIGNMENT_GRADE_COLUMN)

                name_cell.value = assignment_name
                weight_cell.value = weight
                grade_cell.value = 100 if not self.POINTS else weight

                row += 1
            
            #Final Percentage
            MAX_ASSIGNMENT_ROW = row

            final_percentage_formula = ""

            if self.POINTS:
                # Can just be SUM(D:D) / SUM(C:C)
                final_percentage_formula = f"=SUM({ASSIGNMENT_GRADE_COL_LETTER}:{ASSIGNMENT_GRADE_COL_LETTER}) / SUM({ASSIGNMENT_WEIGHT_COL_LETTER}:{ASSIGNMENT_WEIGHT_COL_LETTER})*100"
            else:
                #Final formula needs to look similar to "=SUM(A1*B2, A2*B2, ...)"
                final_percentage_formula = "=SUM(" 
                first = True
                for row in range(STARTING_ROW, MAX_ASSIGNMENT_ROW):
                    weight_cell = sheet.cell(row, ASSIGNMENT_WEIGHT_COLUMN)
                    grade_cell = sheet.cell(row, ASSIGNMENT_GRADE_COLUMN)

                    weight_cell_coord = weight_cell.coordinate
                    grade_cell_coord = grade_cell.coordinate
                    if first:
                        final_percentage_formula += f'{weight_cell_coord}*{grade_cell_coord}'
                        first = not first
                    else:
                        final_percentage_formula += f',{weight_cell_coord}*{grade_cell_coord}'

                final_percentage_formula += ")"
            
            FINAL_PERCENTAGE_CELL = sheet.cell(FINAL_PERCENTAGE_ROW, FINAL_PERCENTAGE_COLUMN)
            FINAL_PERCENTAGE_CELL.value = final_percentage_formula

            if HEADERS:
                cell = sheet.cell(FINAL_PERCENTAGE_ROW, FINAL_PERCENTAGE_COLUMN - 1)
                cell.value = "Final Percentage"
            
            #Final Letter grade
            FINAL_LETTER_CELL = sheet.cell(FINAL_LETTER_ROW, FINAL_LETTER_COLUMN)

            #Final formula needs to look similar to "=IF(A1>F1, A, IF(A1>F2, B, IF(A1>F3, C, ....))"
            final_letter_formula = f"=IF("
            paren_counter = 0
            for letter in sorted(LETTER_GRADES_DICT):
                minval = LETTER_GRADES_DICT[letter]
                final_letter_formula += f'{FINAL_PERCENTAGE_CELL.coordinate} >= {minval}, "{letter}", IF('
                paren_counter += 1

            final_letter_formula = final_letter_formula[0:len(final_letter_formula)-5]
            final_letter_formula += (")"*paren_counter)
            FINAL_LETTER_CELL.value = final_letter_formula


            # Headers
            if HEADERS:
                cell = sheet.cell(FINAL_PERCENTAGE_ROW, FINAL_PERCENTAGE_COLUMN - 1)
                cell.value = "Final Percentage"

                cell = sheet.cell(FINAL_LETTER_ROW, FINAL_LETTER_COLUMN - 1)
                cell.value = "Final Letter Grade"

                cell = sheet.cell(HEADER_ROW, ASSIGNMENT_NAME_COLUMN)
                cell.value = "Assignment Name"

                cell = sheet.cell(HEADER_ROW, ASSIGNMENT_WEIGHT_COLUMN)
                cell.value = "Weight" if not POINTS else "Point Value"
                
                cell = sheet.cell(HEADER_ROW, ASSIGNMENT_GRADE_COLUMN)
                cell.value = "Grade" if not POINTS else "Points Earned"

                cell = sheet.cell(HEADER_ROW, LETTER_GRADE_LETTER_COLUMN)
                cell.value = "Letter Grade"

                cell = sheet.cell(HEADER_ROW, LETTER_GRADE_MINIMUM_COLUMN)
                cell.value = "Minimum Percentage"

            # Save file.
            if not output_name.endswith(".xlsx"): output_name += ".xlsx"
            wbook.save(output_name)
            
        else:
            self.write_to_text_area("Enter the filename for your rubric.")
            self.step_flag = not self.step_flag

    def next(self):
        match self.step:
            case Step.START:
                self.start()
            case Step.GRADES:
                self.handle_grades()
            case Step.GRADE_MINIMUMS:
                self.handle_grade_minimums()
            case Step.ASSIGNMENTS:
                self.handle_assignments()
            case Step.SUBMIT:
                self.handle_submit()
            case _:
                showerror("Something Went Wrong", "Something went wrong. Please report this error on GitHub. ERROR CODE 11")


def main():
    root = Tk()
    RubricCreator(root).pack(side="top", fill='both', expand=True)
    root.mainloop()

if __name__ == "__main__":
    root = Tk()
    rubric_creator = RubricCreator(root)
    rubric_creator.pack(side='top', fill='both', expand=True)
    root.after(100, rubric_creator.next())
    root.mainloop()